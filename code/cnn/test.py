'''
Created on 2018-9-11

@author: 74510
'''
import os
from PIL import Image
import tensorflow as tf
import numpy as np
import matplotlib.pyplot as plt
import get_batch
from alexnet import *
from openpyxl import workbook
import time
# from imageenhance import *
# def get_one_image(img_dir):
#     image = Image.open(img_dir).convert("RGB")
#      
#     image.resize((224,224))
# #     image.show()
#     image_arr = np.array(image)
#     return image_arr

def test(test_file):
    log_dir = 'log/'
#     image_arr = get_one_image(test_file)
    image=Image.open(test_file).convert("RGB")
#     image=tf.cast(image,tf.string)
    max_index=''
    with tf.Graph().as_default():
#         image = tf.cast(image_arr, tf.float32)
#         image=tf.image.resize_image_with_crop_or_pad(image,224,224)
#         image=tf.image.decode_jpeg(image,channels=3)
        image=tf.image.resize_images(image,[224,224])
        image = tf.image.per_image_standardization(image)
        image = tf.reshape(image, [1,224, 224, 3])
        print(image.shape)
        conv1,conv2,conv3,conv4,conv5,pool1,pool2,pool3= inference(image,1,5,train=False)
#         features=inference(image,1,5,train=False)
#         logits = tf.nn.softmax(p)
        x = tf.placeholder(tf.float32,shape = [224,224,3])
        saver = tf.train.Saver()
        with tf.Session() as sess:
            ckpt = tf.train.get_checkpoint_state(log_dir)
            if ckpt and ckpt.model_checkpoint_path:
                global_step = ckpt.model_checkpoint_path.split('/')[-1].split('-')[-1]
                saver.restore(sess, ckpt.model_checkpoint_path)
                print('Loading success')
            else:
                print('No checkpoint')
#             prediction=sess.run(logits)
            feature_conv1=sess.run(conv1)
            feature_conv2=sess.run(conv2)
            feature_conv3=sess.run(conv3)
            feature_conv4=sess.run(conv4)
            feature_conv5=sess.run(conv5)
            feature_pool1=sess.run(pool1)
            feature_pool2=sess.run(pool2)
            feature_pool3=sess.run(pool3)
           
#             max_index = np.argmax(prediction) #np.argmax(a)返回a的最大索引
#             print('预测的标签为：')
#             print(max_index)
#             print('预测的结果为：')
#             if max_index==0:
#                 print('This is the number 0' )
#             elif max_index == 1:
#                 print('This is the number 1' )
#             elif max_index == 2:
#                 print('This is the number 2' )
#             elif max_index == 3:
#                 print('This is the number 3' )
#             elif max_index == 4:
#                 print('This is the number 4' )
#             elif max_index == 5:
#                 print('This is the number 5' )
#             elif max_index == 6:
#                 print('This is the number 6' )
#             elif max_index == 7:
#                 print('This is the number 7' )
#             elif max_index == 8:
#                 print('This is the number 8' )
#             else :
#                 print('This is the number 9' )
    return feature_conv1,feature_conv2,feature_conv3,feature_conv4,feature_conv5,feature_pool1,feature_pool2,feature_pool3
# converimage(feature,image_dir,j)
def converimage(feature,image_dir,j):
    print(type(feature))
    feature1=0
    dim=feature.shape[3]
    print(dim)
    for i in range(dim):
        feature1+=feature[0][:,:,i]
        print(feature1.shape)
#     vis_square(feature1, padsize=1, padval=0)
    image = Image.fromarray(feature1)
#     image = image.convert('RGB')
#     image.save(image_dir+'\{}.jpg'.format(j))
    plt.imshow(image,cmap=plt.cm.jet)
    plt.imsave(image_dir+'\{}.png'.format(j),image,cmap=plt.cm.jet)
#     plt.savefig()
    plt.pause(1)
    plt.close()
# def get_row_col(num_pic):
#     squr = num_pic ** 0.5
#     row = round(squr)
#     col = row + 1 if squr - row > 0 else row
#     return row, col
# def converimage(feature):
#     print(type(feature))
#     feature_map_combination = []
#     dim=feature.shape[3]
#     plt.figure()
#  
#     num_pic = feature.shape[2]
#     row, col = get_row_col(num_pic)
#     print(dim)
#     for i in range(dim):
# #         feature1+=feature[0][:,:,i]
#         feature_map_combination.append(feature[0][:,:,i])
# #         plt.subplot(row, col, i + 1)
# #         print(feature1.shape)
# #         plt.imshow(feature[0][:,:,i])
# #         plt.show()
# #     vis_square(feature1, padsize=1, padval=0)
# #     image = Image.fromarray(feature1)
# #     image = image.convert('RGB')
# #     image.save(image_dir+'\{}.jpg'.format(j))
#     feature_map_sum = sum(ele for ele in feature_map_combination)
#     plt.imshow(feature_map_sum)
#     plt.show()
# #     image.show()
# #     image.close()
# 
image_dir=r'F:\peach\images'
# # feature_conv1,feature_conv2,feature_conv3,feature_conv4,feature_conv5=test(image_dir)
conv1,conv2,conv3,conv4,conv5,pool1,pool2,pool3=test(os.path.join(image_dir,'IMG_5794.JPG'))
# # print(feature_conv1[0].shape)
# # image = Image.fromarray(feature_conv1[0][:,:,3])
# # plt.imshow(image)
# # plt.show()
# # image = image.convert('RGB')
# 
# # vis_square(feature_conv1.transpose(0, 2, 3, 1))
# # data = np.reshape(feature_pool1,())
# # print(feature_pool3)
converimage(conv3,image_dir,3)
# for dir in os.listdir(image_dir):
#     for im_dir in os.listdir(os.path.join(image_dir,dir)):
#         path=os.path.join(image_dir,dir,im_dir)
#         feature_conv1,feature_pool1,feature_conv2,feature_pool2,feature_conv3,feature_conv4,feature_conv5,feature_pool3=test(path)
#         feature_list=[feature_conv1,feature_pool1,feature_conv2,feature_pool2,feature_conv3,feature_conv4,feature_conv5,feature_pool3]
#         for i,fea in enumerate(feature_list):
#             converimage(fea,os.path.join(image_dir,dir),i+1)
        
# print(1,feature_conv1.shape,feature_conv1)
# print(2,feature_conv2.shape,feature_conv2)
# print(3,feature_conv3.shape,feature_conv3)
# print(4,feature_conv4.shape,feature_conv4)
# print(5,feature_conv5.shape,feature_conv5)
# print(1,feature_pool1.shape,feature_pool1)
# print(2,feature_pool2.shape,feature_pool2)
# print(3,feature_pool3.shape,feature_pool3)
# def saveFeatureConv(feature,num,type):
#     attr=['layer']
#     wb=workbook.Workbook()
#     ws=wb.active
#     featureShape=feature.shape
#     layer=featureShape[3]
#     for i in range(featureShape[1]):
#         attr.append('feature{}'.format(i+1))
# #     print(attr)
#     ws.append(attr)
# #     print(feature[0].shape)
# #     print(feature[0][:,:,0].shape,feature[0][:,:,0])
#     for i in range(layer):
#         item=feature[0][:,:,i]
#         print(item)
#         for ie in item:
#             im=[i+1]
#             for it in ie.tolist():
#                 im.append(it)
#             ws.append(im)
#            
#     wb.save(type+'{}.xlsx'.format(num))
#     print('end.............')
# image_dir='../trainmel'
# for dir in os.listdir(image_dir):
#      for im_dir in os.listdir(os.path.join(image_dir,dir)):
#          path=os.path.join(image_dir,dir,im_dir)
#          feature=test(path)
#          saveFeatureConv(feature,3,'pool')
         
# saveFeatureConv(feature,3,pool)
# saveFeatureConv(feature_pool2,2)
# saveFeatureConv(feature_pool3,3)
# saveFeatureConv(feature_conv4,4)
# saveFeatureConv(feature_conv5,5)
# saveFeatureConv(feature_pool3,3)
# def saveimage(image):
#     wb=workbook.Workbook()
#     ws=wb.active
#     attr=[]
#     for i in range(1,9217):
#         attr.append('feature{}'.format(i))
#     ws.append(attr)
#     for img in image:
#         feature=test(img)
#         print(feature)
#         feature=np.array(feature[0]).tolist()
#         print(feature)
# #     for item in feature:
# #         fea.append(item)
# #     features.append(fea)
#         ws.append(feature)
#       
#     wb.save('features-melte.xlsx')
# def savelabel(label):
#        
#     wb=workbook.Workbook()
#     ws=wb.active
#     ws.append(['label'])
#     for la in label:
#         ws.append([la])
#         print('label',la)
#     wb.save('labels-melte.xlsx')
# image,label=get_batch.get_files('../testmel')
# print(image,label)
# saveimage(image)
# savelabel(label)
# print('end...................')
# if __name__ == '__main__':
#     test_dir="../testmel"
#     index=0
#     current_num=0
#     total=0.0#当前文件夹样本总量
#     sess=tf.Session()
#     for file in os.listdir(test_dir):
#         print('begin the test of {0}'.format(file))
#         base_dir=test_dir+'/'+file
#         for image_file in os.listdir(base_dir):
#             #测试数据
#             total+=1
#             image_file=base_dir+"/"+image_file
#             print(image_file)
#             begin_time=time.time()
#             num= test(image_file)
#             print(num)
#             if num==index: current_num+=1
#             current_num=tf.cast(current_num,tf.float32)
#             accury=current_num/total
#             print('the current of accurary:',sess.run(accury),'time',time.time()-begin_time)
#             image=image=tf.gfile.FastGFile(image_file,'rb').read()
#             image=tf.image.decode_jpeg(image)
#             plt.imshow(image.eval(session=sess))
#             plt.pause(1)
#             # 关闭当前显示的图像
#             plt.close()
#         index+=1  
#         if index<=9:
#             print("begin the next type file")
#         else:
#             print('over.......')
#     sess.close()       